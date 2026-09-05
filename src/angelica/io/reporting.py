from __future__ import annotations

from pathlib import Path


def export_solve_result_csv(result, output_path: str) -> None:
    import csv

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    has_temperatures = bool(result.node_temperatures_c)
    has_component_temps = any(
        cf.temperature_in_c is not None for cf in result.component_flows
    )
    has_vf = any(
        getattr(cf, "vapor_fraction", None) is not None for cf in result.component_flows
    )
    has_fw = any(
        getattr(cf, "free_water_fraction", 0.0) > 0.0 for cf in result.component_flows
    )

    with output.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["case", result.case_name])
        writer.writerow(["converged", result.converged])
        writer.writerow([])
        if has_temperatures:
            writer.writerow(["Node", "Pressure (Pa)", "Pressure (kPa)", "Temperature (°C)"])
            for node_id in sorted(result.node_pressures_pa):
                pressure_pa = result.node_pressures_pa[node_id]
                temp_c = result.node_temperatures_c.get(node_id)
                writer.writerow([
                    node_id,
                    round(pressure_pa, 2),
                    round(pressure_pa / 1000.0, 4),
                    round(temp_c, 4) if temp_c is not None else "",
                ])
        else:
            writer.writerow(["Node", "Pressure (Pa)", "Pressure (kPa)"])
            for node_id in sorted(result.node_pressures_pa):
                pressure_pa = result.node_pressures_pa[node_id]
                writer.writerow([node_id, round(pressure_pa, 2), round(pressure_pa / 1000.0, 4)])
        writer.writerow([])
        has_mw = bool(getattr(result, "component_mws", ()))
        header = ["Component", "Mass flow (kg/s)", "Vol. flow (m^3/h)"]
        if has_component_temps:
            header += ["T_in (°C)", "T_out (°C)"]
        if has_vf:
            header.append("Vapor fraction (-)")
        if has_fw:
            header.append("Free water (mol frac)")
        if has_mw:
            header.append("MW_mix (g/mol)")
        writer.writerow(header)
        for component in result.component_flows:
            row = [
                component.label,
                round(component.mass_flow_kg_per_s, 4),
                round(component.volumetric_flow_m3_per_h, 4),
            ]
            if has_component_temps:
                row.append(round(component.temperature_in_c, 4) if component.temperature_in_c is not None else "")
                row.append(round(component.temperature_out_c, 4) if component.temperature_out_c is not None else "")
            if has_vf:
                vf = getattr(component, "vapor_fraction", None)
                row.append(round(vf, 6) if vf is not None else "")
            if has_fw:
                row.append(round(getattr(component, "free_water_fraction", 0.0), 6))
            if has_mw:
                row.append(_mw_mix(getattr(component, "zs", ()), result.component_mws))
            writer.writerow(row)
        _write_balance_rows_csv(writer, result)
        _write_compositions_csv(writer, result)
        _write_component_mass_flows_csv(writer, result)
        _write_phase_report_csv(writer, result)


def export_solve_result_workbook(result, output_path: str) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as _err:
        raise ImportError(
            "Excel export requires the 'openpyxl' package. "
            "Install it with: pip install 'angelica[excel]'"
        ) from _err

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    has_temperatures = bool(result.node_temperatures_c)
    has_component_temps = any(
        cf.temperature_in_c is not None for cf in result.component_flows
    )

    workbook = Workbook()

    pressures_sheet = workbook.active
    pressures_sheet.title = "Pressures"
    if has_temperatures:
        pressures_sheet.append(["Node", "Pressure (Pa)", "Pressure (kPa)", "Temperature (°C)"])
        for node_id in sorted(result.node_pressures_pa):
            pressure_pa = result.node_pressures_pa[node_id]
            temp_c = result.node_temperatures_c.get(node_id)
            pressures_sheet.append([
                node_id,
                round(pressure_pa, 2),
                round(pressure_pa / 1000.0, 4),
                round(temp_c, 4) if temp_c is not None else None,
            ])
    else:
        pressures_sheet.append(["Node", "Pressure (Pa)", "Pressure (kPa)"])
        for node_id in sorted(result.node_pressures_pa):
            pressure_pa = result.node_pressures_pa[node_id]
            pressures_sheet.append([node_id, round(pressure_pa, 2), round(pressure_pa / 1000.0, 4)])

    has_mw = bool(getattr(result, "component_mws", ()))
    has_vf = any(
        getattr(cf, "vapor_fraction", None) is not None for cf in result.component_flows
    )
    has_fw = any(
        getattr(cf, "free_water_fraction", 0.0) > 0.0 for cf in result.component_flows
    )
    flows_sheet = workbook.create_sheet("Flows")
    header = ["Component", "Mass flow (kg/s)", "Vol. flow (m^3/h)"]
    if has_component_temps:
        header += ["T_in (°C)", "T_out (°C)"]
    if has_vf:
        header.append("Vapor fraction (-)")
    if has_fw:
        header.append("Free water (mol frac)")
    if has_mw:
        header.append("MW_mix (g/mol)")
    flows_sheet.append(header)
    for component in result.component_flows:
        row = [
            component.label,
            round(component.mass_flow_kg_per_s, 4),
            round(component.volumetric_flow_m3_per_h, 4),
        ]
        if has_component_temps:
            row.append(round(component.temperature_in_c, 4) if component.temperature_in_c is not None else None)
            row.append(round(component.temperature_out_c, 4) if component.temperature_out_c is not None else None)
        if has_vf:
            vf = getattr(component, "vapor_fraction", None)
            row.append(round(vf, 6) if vf is not None else None)
        if has_fw:
            row.append(round(getattr(component, "free_water_fraction", 0.0), 6))
        if has_mw:
            row.append(_mw_mix(getattr(component, "zs", ()), result.component_mws))
        flows_sheet.append(row)

    _write_balance_sheet(workbook, result)
    _write_compositions_sheet(workbook, result)
    _write_component_mass_flows_sheet(workbook, result)
    _write_phase_report_sheet(workbook, result)
    workbook.save(output)


def _mw_mix(zs: tuple, mws: tuple) -> "float | str":
    """Mixture molecular weight (g/mol) from mole fractions and component MWs."""
    if not zs or not mws:
        return ""
    return round(sum(z * mw_i for z, mw_i in zip(zs, mws)), 4)


def _species_mass_flows(mdot: float, zs: tuple, mws: tuple) -> list:
    """Per-species mass flow rates (kg/s): ṁᵢ = ṁ_total × zᵢ × MWᵢ / MW_mix."""
    mw_mix = sum(z * mw_i for z, mw_i in zip(zs, mws))
    if mw_mix <= 0.0:
        return [0.0] * len(zs)
    return [round(mdot * z * mw_i / mw_mix, 6) for z, mw_i in zip(zs, mws)]


def _write_component_mass_flows_csv(writer, result) -> None:
    names = getattr(result, "component_names", ())
    mws   = getattr(result, "component_mws", ())
    if not names or not mws:
        return
    has_pipe = any(getattr(cf, "zs", ()) for cf in result.component_flows)
    if not has_pipe:
        return
    writer.writerow([])
    writer.writerow(["Component mass flows (kg/s)"])
    writer.writerow(["Pipe", *names])
    for cf in result.component_flows:
        zs = getattr(cf, "zs", ())
        if not zs:
            writer.writerow([cf.label, *[""] * len(names)])
            continue
        writer.writerow([cf.label, *_species_mass_flows(cf.mass_flow_kg_per_s, zs, mws)])


def _write_component_mass_flows_sheet(workbook, result) -> None:
    names = getattr(result, "component_names", ())
    mws   = getattr(result, "component_mws", ())
    if not names or not mws:
        return
    has_pipe = any(getattr(cf, "zs", ()) for cf in result.component_flows)
    if not has_pipe:
        return
    ws = workbook.create_sheet("Component Mass Flows")
    ws.append(["Pipe", *names])
    for cf in result.component_flows:
        zs = getattr(cf, "zs", ())
        if not zs:
            ws.append([cf.label, *[None] * len(names)])
            continue
        ws.append([cf.label, *_species_mass_flows(cf.mass_flow_kg_per_s, zs, mws)])


def _write_compositions_csv(writer, result) -> None:
    names = getattr(result, "component_names", ())
    if not names:
        return
    has_pipe = any(getattr(cf, "zs", ()) for cf in result.component_flows)
    has_node = bool(getattr(result, "node_compositions", {}))
    if not has_pipe and not has_node:
        return
    writer.writerow([])
    writer.writerow(["Compositions (mole fractions)"])
    if has_pipe:
        writer.writerow(["Pipe", *names])
        for cf in result.component_flows:
            zs = getattr(cf, "zs", ())
            writer.writerow([cf.label, *[round(z, 6) for z in zs]])
    if has_node:
        writer.writerow([])
        writer.writerow(["Node", *names])
        for nid in sorted(result.node_compositions):
            zs = result.node_compositions[nid]
            writer.writerow([nid, *[round(z, 6) for z in zs]])


def _write_compositions_sheet(workbook, result) -> None:
    names = getattr(result, "component_names", ())
    if not names:
        return
    has_pipe = any(getattr(cf, "zs", ()) for cf in result.component_flows)
    has_node = bool(getattr(result, "node_compositions", {}))
    if not has_pipe and not has_node:
        return
    ws = workbook.create_sheet("Compositions")
    if has_pipe:
        ws.append(["Pipe", *names])
        for cf in result.component_flows:
            zs = getattr(cf, "zs", ())
            ws.append([cf.label, *[round(z, 6) for z in zs]])
    if has_node:
        ws.append([])
        ws.append(["Node", *names])
        for nid in sorted(result.node_compositions):
            zs = result.node_compositions[nid]
            ws.append([nid, *[round(z, 6) for z in zs]])


def _write_balance_rows_csv(writer, result) -> None:
    gb = result.global_balance
    geb = result.global_energy_balance
    if gb is None and geb is None:
        return
    writer.writerow([])
    writer.writerow(["Global Balance"])
    if gb is not None:
        writer.writerow(["Quantity", "Value", "Unit"])
        writer.writerow(["Mass flow in",  round(gb.mass_inlet_kg_per_s, 6),  "kg/s"])
        writer.writerow(["Mass flow out", round(gb.mass_outlet_kg_per_s, 6), "kg/s"])
        writer.writerow(["Mass balance error", round(gb.mass_error_pct, 6),  "%"])
    if geb is not None:
        writer.writerow(["Enthalpy in",       round(geb.enthalpy_in_kw, 6),    "kW"])
        writer.writerow(["Enthalpy out",      round(geb.enthalpy_out_kw, 6),   "kW"])
        writer.writerow(["Heat sources",      round(geb.heat_sources_kw, 6),   "kW"])
        writer.writerow(["Heat wall loss",    round(geb.heat_wall_loss_kw, 6), "kW"])
        writer.writerow(["Energy balance error (kW)", round(geb.energy_error_kw, 9), "kW"])
        writer.writerow(["Energy balance error (%)",  round(geb.energy_error_pct, 6), "%"])


def _write_balance_sheet(workbook, result) -> None:
    gb = result.global_balance
    geb = result.global_energy_balance
    if gb is None and geb is None:
        return
    ws = workbook.create_sheet("Balance")
    ws.append(["Quantity", "Value", "Unit"])
    if gb is not None:
        ws.append(["Mass flow in",  round(gb.mass_inlet_kg_per_s, 6),  "kg/s"])
        ws.append(["Mass flow out", round(gb.mass_outlet_kg_per_s, 6), "kg/s"])
        ws.append(["Mass balance error", round(gb.mass_error_pct, 6),  "%"])
    if geb is not None:
        ws.append(["Enthalpy in",       round(geb.enthalpy_in_kw, 6),    "kW"])
        ws.append(["Enthalpy out",      round(geb.enthalpy_out_kw, 6),   "kW"])
        ws.append(["Heat sources",      round(geb.heat_sources_kw, 6),   "kW"])
        ws.append(["Heat wall loss",    round(geb.heat_wall_loss_kw, 6), "kW"])
        ws.append(["Energy balance error (kW)", round(geb.energy_error_kw, 9), "kW"])
        ws.append(["Energy balance error (%)",  round(geb.energy_error_pct, 6), "%"])


def print_solve_result(result, detailed: bool = False) -> None:
    print(f"Case: {result.case_name}")
    print(f"Converged: {result.converged}")
    print("")
    _print_iteration_summary(result)
    print("")
    _print_node_table(result.node_pressures_pa, result.node_temperatures_c or None)
    print("")
    _print_component_table(result.component_flows)
    if result.turbulent_metrics:
        print("")
        _print_final_turbulent_metrics(result.turbulent_metrics[-1])
    if detailed:
        print("")
        _print_history_section("Laminar Initialisation History", result.laminar_history)
        if result.laminar_metrics:
            print("")
            _print_iteration_metrics_history("Laminar Iteration Metrics", result.laminar_metrics)
        print("")
        _print_history_section("Turbulent Correction History", result.turbulent_history)
        if result.turbulent_metrics:
            print("")
            _print_iteration_metrics_history("Turbulent Iteration Metrics", result.turbulent_metrics)


def print_detailed_solve_result(result) -> None:
    print_solve_result(result, detailed=True)


def save_convergence_plot(result, output_path: str) -> None:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    fig, axes = plt.subplots(2, 1, figsize=(9, 7), sharex=False)
    _plot_history(axes[0], "Laminar correction history", result.laminar_history)
    _plot_history(axes[1], "Turbulent correction history", result.turbulent_history)
    fig.tight_layout()
    fig.savefig(output, dpi=160)
    plt.close(fig)


def _print_node_table(node_pressures_pa, node_temperatures_c=None) -> None:
    print("Node Pressures")
    if node_temperatures_c:
        print(f"{'Node':>6}  {'Pressure (Pa)':>14}  {'Pressure (kPa)':>14}  {'Temp (°C)':>10}")
        for node_id in sorted(node_pressures_pa):
            pressure_pa = node_pressures_pa[node_id]
            temp = node_temperatures_c.get(node_id)
            temp_str = f"{temp:>10.4f}" if temp is not None else f"{'—':>10}"
            print(f"{node_id:>6}  {pressure_pa:>14.2f}  {pressure_pa / 1000.0:>14.4f}  {temp_str}")
    else:
        print(f"{'Node':>6}  {'Pressure (Pa)':>14}  {'Pressure (kPa)':>14}")
        for node_id in sorted(node_pressures_pa):
            pressure_pa = node_pressures_pa[node_id]
            print(f"{node_id:>6}  {pressure_pa:>14.2f}  {pressure_pa / 1000.0:>14.4f}")


def _print_component_table(component_flows) -> None:
    print("Component Flows")
    has_temps = any(cf.temperature_in_c is not None for cf in component_flows)
    if has_temps:
        print(f"{'Component':<20}  {'Mass flow (kg/s)':>16}  {'Vol. flow (m^3/h)':>18}  {'T_in (°C)':>10}  {'T_out (°C)':>10}")
        for component in component_flows:
            t_in = f"{component.temperature_in_c:>10.4f}" if component.temperature_in_c is not None else f"{'—':>10}"
            t_out = f"{component.temperature_out_c:>10.4f}" if component.temperature_out_c is not None else f"{'—':>10}"
            print(
                f"{component.label:<20}  "
                f"{component.mass_flow_kg_per_s:>16.4f}  "
                f"{component.volumetric_flow_m3_per_h:>18.4f}  "
                f"{t_in}  {t_out}"
            )
    else:
        print(f"{'Component':<20}  {'Mass flow (kg/s)':>16}  {'Vol. flow (m^3/h)':>18}")
        for component in component_flows:
            print(
                f"{component.label:<20}  "
                f"{component.mass_flow_kg_per_s:>16.4f}  "
                f"{component.volumetric_flow_m3_per_h:>18.4f}"
            )


def _print_iteration_summary(result) -> None:
    print("Iteration Summary")
    print(f"  Laminar initialisation iterations: {len(result.laminar_history)}")
    print(f"  Turbulent iterations:              {len(result.turbulent_history)}")


def _print_history_section(title: str, history) -> None:
    print(title)
    print(_ascii_history_plot(history))
    print(f"{'Iter':>6}  {'Correction':>14}")
    for iteration, value in enumerate(history, start=1):
        print(f"{iteration:>6}  {value:>14.9f}")


def _print_final_turbulent_metrics(metrics) -> None:
    print("Final Turbulent Metrics")
    print(f"  Max abs pressure correction (Pa): {metrics.pressure_correction_abs_pa:.9f}")
    print(f"  Max rel pressure correction:      {metrics.pressure_correction_rel:.9e}")
    print(f"  Max nodal mass imbalance (−):     {metrics.max_nodal_mass_imbalance_rel:.9e}")


def _print_iteration_metrics_history(title: str, metrics_history) -> None:
    print(title)
    print(
        f"{'Iter':>6}  "
        f"{'Abs Corr. (Pa)':>16}  "
        f"{'Rel Corr.':>12}  "
        f"{'Max Mass Imb. (−)':>20}"
    )
    for iteration, metrics in enumerate(metrics_history, start=1):
        print(
            f"{iteration:>6}  "
            f"{metrics.pressure_correction_abs_pa:>16.9f}  "
            f"{metrics.pressure_correction_rel:>12.5e}  "
            f"{metrics.max_nodal_mass_imbalance_rel:>20.9e}"
        )


def _ascii_history_plot(history, width: int = 48) -> str:
    if not history:
        return "(no data)"

    blocks = " .:-=+*#%@"
    minimum = min(history)
    maximum = max(history)
    span = maximum - minimum
    if span == 0:
        line = blocks[-1] * len(history)
    else:
        indices = []
        for value in history:
            normalised = (value - minimum) / span
            indices.append(int(round(normalised * (len(blocks) - 1))))
        line = "".join(blocks[index] for index in indices)

    if len(line) > width:
        step = len(line) / float(width)
        compressed = []
        for idx in range(width):
            compressed.append(line[int(idx * step)])
        line = "".join(compressed)

    return f"[{line}]  min={minimum:.6g} max={maximum:.6g}"


def _plot_history(axis, title: str, history) -> None:
    iterations = list(range(1, len(history) + 1))
    axis.plot(iterations, history, marker="o", linewidth=1.5)
    axis.set_title(title)
    axis.set_xlabel("Iteration")
    axis.set_ylabel("Correction")
    axis.grid(True, alpha=0.3)


# ── Phase Report helpers ──────────────────────────────────────────────────────

_WATER_IDS = {"water", "h2o", "7732-18-5"}

_THRESH = 1e-4  # minimum mole fraction to consider a phase present


def _hc_component_names(result) -> tuple[str, ...]:
    """Return component names with water removed (HC-only basis for phase compositions)."""
    names = getattr(result, "component_names", ())
    return tuple(n for n in names if n.lower() not in _WATER_IDS)


def _has_phase_detail(result) -> bool:
    """True when any pipe has EOS phase data (VF is enough — available since compositional solver)."""
    return any(
        getattr(cf, "vapor_fraction", None) is not None
        for cf in result.component_flows
    )


def _phase_state_label(vf, lf, fw) -> str:
    """Plain-language phase state for one pipe.

    vf  = gas mole fraction of feed (HC gas + water vapour)
    lf  = HC liquid mole fraction of feed (None if not computed)
    fw  = free liquid water mole fraction of feed
    """
    has_gas = vf is not None and vf > _THRESH
    has_liq = lf is not None and lf > _THRESH
    has_fw  = fw > _THRESH

    if not has_gas and not has_liq and not has_fw:
        # everything accounted for as liquid HC (VF≈0, no water)
        return "Single-phase liquid"

    if has_gas and not has_liq and not has_fw:
        return "Single-phase gas"

    phases = []
    if has_gas:
        phases.append("Gas")
    if has_liq:
        phases.append("HC liquid")
    if has_fw:
        phases.append("Free water")

    n = len(phases)
    label = " + ".join(phases)
    return f"{n} phases: {label}"


def _hc_zs_for(cf, all_names: tuple) -> tuple:
    """HC feed composition (water removed, renormalised) from a ComponentFlowResult."""
    zs = getattr(cf, "zs", ())
    if not zs or not all_names:
        return ()
    hc_pairs = [(z, n) for z, n in zip(zs, all_names) if n.lower() not in _WATER_IDS]
    if not hc_pairs:
        return ()
    total = sum(z for z, _ in hc_pairs)
    if total <= 0:
        return ()
    return tuple(z / total for z, _ in hc_pairs)


def _write_phase_report_csv(writer, result) -> None:
    if not _has_phase_detail(result):
        return
    hc_names   = _hc_component_names(result)
    all_names  = getattr(result, "component_names", ())
    has_liq_phase = any(
        (getattr(cf, "liquid_fraction", None) or 0.0) > _THRESH
        for cf in result.component_flows
    )

    writer.writerow([])
    writer.writerow(["=== PHASE REPORT ==="])
    writer.writerow(["V + L + Free water = 1  (mole fractions of total feed)"])

    # Phase fractions per pipe
    writer.writerow([])
    writer.writerow(["Pipe", "Phase state", "V  gas (-)", "L  HC liquid (-)", "Free water (-)"])
    for cf in result.component_flows:
        vf = getattr(cf, "vapor_fraction", None)
        lf = getattr(cf, "liquid_fraction", None)
        fw = getattr(cf, "free_water_fraction", 0.0)
        writer.writerow([
            cf.label,
            _phase_state_label(vf, lf, fw),
            round(vf, 6) if vf is not None else "",
            round(lf, 6) if lf is not None else "",
            round(fw, 6),
        ])

    # Composition table: z_i, y_i, and x_i (only when HC liquid exists)
    has_gas_zs = any(getattr(cf, "gas_phase_zs", ()) for cf in result.component_flows)
    if has_gas_zs and hc_names:
        writer.writerow([])
        if has_liq_phase:
            writer.writerow(["Phase compositions  (mole fractions within each phase)"])
            writer.writerow(["Pipe", "Component", "z_i  feed", "y_i  gas", "x_i  HC liquid"])
        else:
            writer.writerow(["Gas phase composition  y_i  (mole fractions within the gas phase)"])
            writer.writerow(["Pipe", "Component", "z_i  feed", "y_i  gas"])

        for cf in result.component_flows:
            y    = getattr(cf, "gas_phase_zs", ())
            x    = getattr(cf, "liquid_phase_zs", ()) if has_liq_phase else ()
            z_hc = _hc_zs_for(cf, all_names)
            for idx, comp in enumerate(hc_names):
                zi = round(z_hc[idx], 6) if idx < len(z_hc) else ""
                yi = round(y[idx],    6) if idx < len(y)    else ""
                if has_liq_phase:
                    xi = round(x[idx], 6) if idx < len(x) else ""
                    writer.writerow([cf.label if idx == 0 else "", comp, zi, yi, xi])
                else:
                    writer.writerow([cf.label if idx == 0 else "", comp, zi, yi])


def _write_phase_report_sheet(workbook, result) -> None:
    if not _has_phase_detail(result):
        return
    hc_names   = _hc_component_names(result)
    all_names  = getattr(result, "component_names", ())
    has_liq_phase = any(
        (getattr(cf, "liquid_fraction", None) or 0.0) > _THRESH
        for cf in result.component_flows
    )

    ws = workbook.create_sheet("Phase Report")
    ws.append(["V + L + Free water = 1  (mole fractions of total feed)"])
    ws.append([])

    # Phase fractions per pipe
    ws.append(["Pipe", "Phase state", "V  gas (-)", "L  HC liquid (-)", "Free water (-)"])
    for cf in result.component_flows:
        vf = getattr(cf, "vapor_fraction", None)
        lf = getattr(cf, "liquid_fraction", None)
        fw = getattr(cf, "free_water_fraction", 0.0)
        ws.append([
            cf.label,
            _phase_state_label(vf, lf, fw),
            round(vf, 6) if vf is not None else None,
            round(lf, 6) if lf is not None else None,
            round(fw, 6),
        ])

    # Composition table
    has_gas_zs = any(getattr(cf, "gas_phase_zs", ()) for cf in result.component_flows)
    if has_gas_zs and hc_names:
        ws.append([])
        if has_liq_phase:
            ws.append(["Phase compositions  (mole fractions within each phase)"])
            ws.append(["Pipe", "Component", "z_i  feed", "y_i  gas", "x_i  HC liquid"])
        else:
            ws.append(["Gas phase composition  y_i  (mole fractions within the gas phase)"])
            ws.append(["Pipe", "Component", "z_i  feed", "y_i  gas"])

        for cf in result.component_flows:
            y    = getattr(cf, "gas_phase_zs", ())
            x    = getattr(cf, "liquid_phase_zs", ()) if has_liq_phase else ()
            z_hc = _hc_zs_for(cf, all_names)
            for idx, comp in enumerate(hc_names):
                zi = round(z_hc[idx], 6) if idx < len(z_hc) else None
                yi = round(y[idx],    6) if idx < len(y)    else None
                if has_liq_phase:
                    xi = round(x[idx], 6) if idx < len(x) else None
                    ws.append([cf.label if idx == 0 else None, comp, zi, yi, xi])
                else:
                    ws.append([cf.label if idx == 0 else None, comp, zi, yi])
